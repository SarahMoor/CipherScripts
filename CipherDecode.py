# -*- coding: utf-8 -*-
import os
import argparse
from Cryptodome.Cipher import AES
from sys import argv
import binascii
from openpyxl import load_workbook

global hex_codes


def GetArgs():
    '''
    parses script arguments to make running this script more intuitive

    Raises
    ------
    FileNotFoundError
        raises error if file cannot be found.

    Returns
    -------
    args : string
        returns the filename of the txt to be encrypted.

    '''
    parser = argparse.ArgumentParser()
    parser.add_argument('-e', '--encryptedfile', metavar='\b',
                        help="name of file to be decrypted", action="store", dest="encrypted_file")
    parser.add_argument('-m', '--monomerfile', metavar='\b',
                        help="name of file with monomer to hex assignments", action="store", dest="monomer_hex_assignment")
    parser.add_argument('-t', '--template', metavar='\b',
                        help="template containing monomer masses", action="store", dest="LCMS_template")
    args = parser.parse_args()

    # Check if file exists
    if os.path.exists('{}'.format(args.encrypted_file)) is False:
        print('{}'.format(args.encrypted_file))
        raise FileNotFoundError('Cannot find {}'.format(args.encrypted_file))
    if os.path.exists('{}'.format(args.monomer_hex_assignment)) is False:
        print('{}'.format(args.monomer_hex_assignment))
        raise FileNotFoundError(
            'Cannot find {}'.format(args.monomer_hex_assignment))
    if os.path.exists('{}'.format(args.LCMS_template)) is False:
        print('{}'.format(args.LCMS_template))
        raise FileNotFoundError('Cannot find {}'.format(args.LCMS_template))

    return args


def MassToHex(value1, value2, hex_codes):
    """
    Match LC/MS Masses to Monomers

    Parameters
    ----------
    value1 : int
        Parent mass.
    value2 : int
        Mass after loss of one unknown monomer.

    Returns
    -------
    hexadecimal character from list : string
        Gets the difference between value1 and value2 to see which monomer was
        lost, then returns monomer that matches this mass difference.

    """
    mass_match = float(value1) - float(value2)
    for i in hex_codes.values():
        if (i + 1.5 >= mass_match and i - 1.5 <= mass_match):
            return list(hex_codes.keys())[list(hex_codes.values()).index(i)]
    print()
    print(mass_match, "NOT MATCHED")
    return "!"


def MakeBitstring(sheet, hex_codes):
    """
    Takes in excel sheet with templated LC/MS data and converts this first to
    it's hexadecimal format, and then to binary

    Parameters
    ----------
    sheet : xlsx reader object
        Templated LC/MS data in an excel sheet.

    Returns
    -------
    encoded_bitstring : string
        hex digits in string format.

    """
    encoded_hexstring = ""
    new_oligomer = 0
    for col in sheet.iter_cols(min_row=2, min_col=2, values_only=True):
        start_oligomer = 0
        oligomer_number = 0
        for entry in col:
            oligomer_number += 1
            if entry == None:
                break
            entry = float(entry)
            new_oligomer = entry
            if start_oligomer == 0:
                start_oligomer = new_oligomer
            elif oligomer_number == 2:
                start_oligomer = new_oligomer
            else:
                hex_value = str(
                    MassToHex(start_oligomer, new_oligomer, hex_codes))[0]
                encoded_hexstring += hex_value
                start_oligomer = new_oligomer
    return encoded_hexstring


def ConvertToFloat(mass_list):
    total_list = []
    for x in mass_list:
        x = float(x)
        total_list.append(x)
    return total_list


def main():
    # get name of csv to read in and write out
    #script, encrypted_file, monomer_hex_assignment, LCMS_template = argv

    args = GetArgs()
    encrypted_file = args.encrypted_file
    monomer_hex_assignment = args.monomer_hex_assignment
    LCMS_template = args.LCMS_template

    # open and parse file to be decrypted
    encrypted_text = open(encrypted_file, "rb")
    nonce, tag, ciphertext = [encrypted_text.read(x) for x in (16, 16, -1)]

    ####### READ IN HEX ASSIGNMENTS TO MONOMER MASSES ######
    codes_workbook = load_workbook(filename=monomer_hex_assignment)
    sheet1 = codes_workbook.active
    hex_codes = {}

    for row in sheet1.iter_rows(min_row=2, values_only=True):
        hex_codes[row[0]] = row[1]

    ####### END READ IN HEX ASSIGNMENTS TO MONOMER MASSES ######

    ####### END READ IN LCMS TEMPLATE WITH MASSES ######
    template_workbook = load_workbook(filename=(LCMS_template))
    LCMSsheet = template_workbook.active

    encoded_bitstring = MakeBitstring(LCMSsheet, hex_codes)
    print("hexadecimal key:")
    print((encoded_bitstring))
    print ()
    
    end_length = len(encoded_bitstring) * 4

    hex_as_int = int(encoded_bitstring, 16)
    hex_as_binary = bin(hex_as_int)
    padded_binary = hex_as_binary[2:].zfill(end_length)
    print("binary key:")
    print(padded_binary)
    print ()

    key = binascii.unhexlify(encoded_bitstring)
    print(".bin format:")
    print(key)
    print ()

    ####### END READ IN LCMS TEMPLATE WITH MASSES ######

    cipher = AES.new(key, AES.MODE_EAX, nonce)

    try:
        data = cipher.decrypt_and_verify(ciphertext, tag)
        data = data.decode('utf-8')
        decrypted_file = open("decrypted_file.txt", 'w+')
        decrypted_file.write(data)
        decrypted_file.close()
        print("Successfully Decrypted")
    except ValueError:
        print("Incorrect Key")


main()
